# src/app_google/get_google.py
"""
Сервис для обработки Google Sheets с поддержкой ограниченного доступа.
Зависимости: src.config.logger, google-auth, google-auth-oauthlib, gspread, openpyxl
"""
import asyncio
import pickle
import gspread

from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook

from src.app_google.config import APP_GOOGLE_FILE
from src.config.logger import logger


class GoogleSheetProcessor:
    """
    Класс для работы с закрытыми Google Таблицами через OAuth 2.0.

    Поддерживает:
    - Авторизацию через браузер (первый запуск)
    - Кэширование токена и контента файла
    - Асинхронный интерфейс с выполнением блокирующих операций в потоке
    """

    # Права доступа: только чтение таблиц
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

    # Пути к файлам авторизации (относительно корня проекта или абсолютные)
    CREDENTIALS_FILE = Path('credentials.json')  # Скачать из Google Cloud Console
    TOKEN_FILE = Path('token_google.pkl')  # Создается автоматически

    def __init__(self, timeout: int = 30, spreadsheet_id: Optional[str] = None):
        """
        Инициализация процессора.

        Args:
            timeout: Таймаут сетевых запросов в секундах
            spreadsheet_id: Идентификатор таблицы (из URL между /d/ и /edit).
                           Если не указан, берётся из APP_GOOGLE_FILE.
        """
        self.timeout = timeout
        self.spreadsheet_id = spreadsheet_id or APP_GOOGLE_FILE.strip()

        # Кэш контента файла
        self._cached_content: bytes | None = None
        self._cached_file_code: str | None = None

        # Клиент gspread (инициализируется при первой авторизации)
        self._client: gspread.Client | None = None

    def _get_authenticated_client(self) -> gspread.Client:
        """
        Получить авторизованный клиент gspread.

        При первом вызове:
        - Проверяет сохранённый токен
        - При необходимости запускает браузер для входа
        - Сохраняет новый токен

        Returns:
            gspread.Client: Авторизованный клиент
        """
        # Если клиент уже создан и токен валиден — возвращаем его
        if self._client is not None:
            try:
                # Проверяем валидность токена (может выбросить исключение)
                self._client.auth.refresh(Request())
                return self._client
            except Exception:
                logger.debug("Токен истёк, выполняем повторную авторизацию...")
                self._client = None

        creds = None

        # Пробуем загрузить сохранённый токен
        if self.TOKEN_FILE.exists():
            try:
                with open(self.TOKEN_FILE, 'rb') as f:
                    creds = pickle.load(f)
            except Exception as e:
                logger.warning(f"Не удалось загрузить токен: {e}. Удаляю повреждённый файл.")
                self.TOKEN_FILE.unlink(missing_ok=True)
                creds = None

        # Если токена нет или он невалиден — запускаем поток авторизации
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                # Пробуем обновить токен без браузера
                try:
                    creds.refresh(Request())
                    logger.info("Токен успешно обновлён")
                except Exception as e:
                    logger.warning(f"Не удалось обновить токен: {e}. Запрашиваю вход через браузер...")
                    creds = None

            if not creds:
                # Требуется вход через браузер
                if not self.CREDENTIALS_FILE.exists():
                    raise FileNotFoundError(
                        f"Файл {self.CREDENTIALS_FILE} не найден!\n"
                        "Скачайте его из Google Cloud Console: "
                        "APIs & Services → Credentials → Create Credentials → OAuth client ID (Desktop app)"
                    )

                logger.info("🔐 Откройте браузер для авторизации в Google...")
                flow = InstalledAppFlow.from_client_secrets_file(
                    str(self.CREDENTIALS_FILE),
                    self.SCOPES
                )
                creds = flow.run_local_server(port=0, open_browser=True)
                logger.info("✅ Авторизация успешна")

            # Сохраняем токен для будущих запусков
            try:
                with open(self.TOKEN_FILE, 'wb') as f:
                    pickle.dump(creds, f)
                self.TOKEN_FILE.chmod(0o600)  # Защита файла токена
                logger.debug(f"Токен сохранён в {self.TOKEN_FILE}")
            except Exception as e:
                logger.warning(f"Не удалось сохранить токен: {e}. При следующем запуске потребуется вход.")

        # Создаём и возвращаем клиент
        self._client = gspread.authorize(creds)
        return self._client

    async def download_file(self, file_code: Optional[str] = None) -> bool:
        """
        Скачать Google Sheet в память и закэшировать.

        При повторном вызове с тем же file_code — скачивание пропускается.

        Важно: Авторизация выполняется от имени пользователя,
        чей аккаунт имеет доступ к таблице.

        Args:
            file_code: Идентификатор файла (опционально, переопределяет значение из __init__).

        Returns:
            bool: True если файл загружен успешно, False при ошибке.
        """
        target_id = (file_code or self.spreadsheet_id).strip()

        # Если файл уже загружен и ID совпадает — используем кэш
        if self._cached_content is not None and self._cached_file_code == target_id:
            logger.debug(f"✅ Используем кэш для файла: {target_id}")
            return True

        try:
            logger.info(f"📥 Загрузка таблицы: {target_id}")

            # Выполняем блокирующий вызов API в отдельном потоке
            def _fetch_sync() -> bytes:
                client = self._get_authenticated_client()
                spreadsheet = client.open_by_key(target_id)
                # Экспортируем в байты XLSX через внутренний метод gspread
                return spreadsheet.export(gspread.utils.ExportFormat.EXCEL)

            content = await asyncio.to_thread(_fetch_sync)

            # Кэшируем результат
            self._cached_content = content
            self._cached_file_code = target_id

            logger.debug(f"✅ Файл загружен и закэширован: {len(content)} байт")
            return True

        except gspread.exceptions.SpreadsheetNotFound:
            logger.error(f"❌ 404: Таблица '{target_id}' не найдена. Проверьте ID и права доступа аккаунта.")
            return False
        except gspread.exceptions.APIError as e:
            logger.error(f"❌ API Error: {e}. Возможно, у аккаунта нет доступа к таблице.")
            return False
        except FileNotFoundError as e:
            logger.error(f"❌ {e}")
            return False
        except Exception as e:
            logger.error(f"❌ download_file: {type(e).__name__}: {e}", exc_info=True)
            return False

    def _parse_sheet_names_sync(self, content: bytes) -> list[str]:
        """Синхронный парсинг имён листов (для выполнения в потоке)."""
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        return workbook.sheetnames

    async def get_sheet_names(self) -> list[str] | None:
        """
        Получить список всех листов из закэшированного файла.

        Returns:
            list[str] | None: Список имён листов или None если файл не загружен/ошибка.
        """
        if self._cached_content is None:
            logger.error("❌ Файл не загружен. Сначала вызовите download_file(file_code)")
            return None

        try:
            sheet_names = await asyncio.to_thread(self._parse_sheet_names_sync, self._cached_content)
            logger.info(f"📋 Найдено листов ({len(sheet_names)}): {sheet_names}")
            return sheet_names
        except Exception as e:
            logger.error(f"❌ Ошибка парсинга имён листов: {e}", exc_info=True)
            return None

    def _parse_sheet_data_sync(self, content: bytes, list_name: str) -> list[dict[str, Any]] | None:
        """Синхронный парсинг данных листа (для выполнения в потоке)."""
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)

        if list_name not in workbook.sheetnames:
            logger.error(f"Лист '{list_name}' не найден. Доступные: {workbook.sheetnames}")
            return None

        worksheet = workbook[list_name]
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            logger.warning("Лист пустой")
            return []

        # Заголовки
        headers = [
            str(h).strip() if h is not None else f"column_{idx}"
            for idx, h in enumerate(rows[0])
        ]

        # Данные
        data = []
        for row in rows[1:]:
            if row and any(cell is not None for cell in row):
                record = {}
                for idx, header in enumerate(headers):
                    value = row[idx] if idx < len(row) else None
                    if isinstance(value, str):
                        value = value.strip()
                    record[header] = value
                data.append(record)

        return data

    async def get_sheet_data(self, list_name: str) -> list[dict[str, Any]] | None:
        """
        Получить данные конкретного листа из закэшированного файла.

        Args:
            list_name: Имя листа для извлечения.

        Returns:
            list[dict] | None: Данные листа как список словарей или None при ошибке.
        """
        if self._cached_content is None:
            logger.error("❌ Файл не загружен. Сначала вызовите download_file(file_code)")
            return None

        try:
            data = await asyncio.to_thread(self._parse_sheet_data_sync, self._cached_content, list_name)
            if data is not None:
                logger.info(f"✅ Получено строк из листа '{list_name}': {len(data)}")
            return data
        except Exception as e:
            logger.error(f"❌ Ошибка парсинга листа '{list_name}': {e}", exc_info=True)
            return None

    async def get_sheet_columns(self, list_name: str) -> list[str] | None:
        """
        Получить только имена столбцов из листа (без загрузки всех данных).

        Args:
            list_name: Имя листа.

        Returns:
            list[str] | None: Список имён столбцов или None при ошибке.
        """
        if self._cached_content is None:
            logger.error("❌ Файл не загружен. Сначала вызовите download_file(file_code)")
            return None

        def _extract_columns_sync(content: bytes, list_name: str) -> list[str] | None:
            try:
                workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)

                if list_name not in workbook.sheetnames:
                    logger.error(f"Лист '{list_name}' не найден. Доступные: {workbook.sheetnames}")
                    return None

                worksheet = workbook[list_name]
                header_row = next(worksheet.iter_rows(values_only=True, max_row=1), None)
                if not header_row:
                    logger.warning("Лист пустой, нет заголовков")
                    return []

                return [
                    str(h).strip() if h is not None else f"column_{idx}"
                    for idx, h in enumerate(header_row)
                ]
            except Exception as e:
                logger.error(f"❌ Ошибка чтения заголовков: {e}", exc_info=True)
                return None

        try:
            columns = await asyncio.to_thread(_extract_columns_sync, self._cached_content, list_name)
            if columns is not None:
                logger.info(f"📊 Столбцы листа '{list_name}': {columns}")
            return columns
        except Exception as e:
            logger.error(f"❌ Ошибка получения столбцов: {e}", exc_info=True)
            return None

    def clear_cache(self) -> None:
        """Очистить кэш файла (освободить память)."""
        self._cached_content = None
        self._cached_file_code = None
        logger.debug("🗑️ Кэш контента очищен")

    def clear_token(self) -> None:
        """Удалить сохранённый токен авторизации (для сброса сессии)."""
        if self.TOKEN_FILE.exists():
            self.TOKEN_FILE.unlink()
            logger.info("🔑 Токен авторизации удалён")
        self._client = None

    @property
    def is_file_loaded(self) -> bool:
        """Проверить, загружен ли файл в кэш."""
        return self._cached_content is not None

    @property
    def cached_file_code(self) -> str | None:
        """Вернуть идентификатор закэшированного файла."""
        return self._cached_file_code

    @property
    def is_authenticated(self) -> bool:
        """Проверить, есть ли активная авторизация."""
        if self._client is None:
            return False
        try:
            return self._client.auth.valid
        except Exception:
            return False
