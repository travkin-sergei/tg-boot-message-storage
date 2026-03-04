import asyncio
import aiohttp
import pyarrow as pa
import pyarrow.parquet as pq

from pathlib import Path
from io import BytesIO
from typing import Optional, List, Any, Dict
from openpyxl import load_workbook

from src.app_log import logger


def _process_workbook_sync(content: bytes,
                           list_name: str) -> Optional[List[Dict[str, Any]]]:
    """
    Синхронная функция для парсинга XLSX в память.
    Выполняется в отдельном потоке.
    """
    try:
        workbook = load_workbook(filename=BytesIO(content))

        if list_name not in workbook.sheetnames:
            logger.error(f"❌ Лист '{list_name}' не найден. Доступные: {workbook.sheetnames}")
            return None

        worksheet = workbook[list_name]
        rows: List[tuple] = list(worksheet.iter_rows(values_only=True))

        if not rows:
            logger.warning("⚠️ Лист пустой")
            return None

        # Заголовки
        headers: List[str] = [
            str(header).strip() if header is not None else f"column_{idx}"
            for idx, header in enumerate(rows[0])
        ]

        # Данные
        data: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if row and any(cell is not None for cell in row):
                record: Dict[str, Any] = {}
                for idx, header in enumerate(headers):
                    value: Any = row[idx] if idx < len(row) else None
                    if isinstance(value, str):
                        value = value.strip()
                    record[header] = value
                data.append(record)

        return data

    except Exception as e:
        logger.error(f"❌ Ошибка при парсинге workbook: {e}", exc_info=True)
        return None


def _save_parquet_sync(data: List[Dict[str, Any]], output_file: Path) -> str:
    """
    Синхронная функция для сохранения в Parquet.
    Выполняется в отдельном потоке.
    """
    table = pa.Table.from_pylist(data)
    pq.write_table(table, output_file, compression='snappy')
    return str(output_file.resolve())


async def google_to_parquet(list_name: str,
                            file_code: str,
                            output_path: str) -> Optional[str]:
    """
    Асинхронно скачивает Google Sheet, парсит и сохраняет в Parquet.

    Args:
        list_name: Название листа в Google таблице.
        file_code: Идентификатор файла (из URL Google Sheets).
        output_path: Полный путь для сохранения .parquet файла.

    Returns:
        str: Путь к сохранённому файлу при успехе.
        None: При любой ошибке.
    """

    link: str = f'https://docs.google.com/spreadsheets/d/{file_code}/export?format=xlsx'

    try:
        logger.info(f"Загрузка Google Sheet: {file_code} / лист '{list_name}'")

        # 1. Асинхронное скачивание через aiohttp
        async with aiohttp.ClientSession() as session:
            async with session.get(link, timeout=aiohttp.ClientTimeout(total=30)) as response:
                if response.status == 404:
                    logger.error(
                        "404: Таблица не найдена. "
                        "Проверьте file_code и доступ 'Все, у кого есть ссылка'"
                    )
                    return None
                response.raise_for_status()
                content: bytes = await response.read()

        # 2. Парсинг XLSX в отдельном потоке (блокирующая операция)
        data: Optional[List[Dict[str, Any]]] = await asyncio.to_thread(
            _process_workbook_sync,
            content,
            list_name
        )

        if data is None:
            return None

        if not data:
            logger.warning("Нет данных для сохранения после парсинга")
            return None

        # 3. Подготовка пути и сохранение в Parquet в отдельном потоке
        output_file = Path(output_path)
        await asyncio.to_thread(output_file.parent.mkdir, parents=True, exist_ok=True)

        result_path: str = await asyncio.to_thread(
            _save_parquet_sync,
            data,
            output_file
        )

        logger.info(f"Сохранено {len(data)} строк в {result_path}")
        return result_path

    except aiohttp.ClientError as error:
        logger.error(f"Ошибка HTTP при загрузке: {error}")
        return None
    except Exception as error:
        logger.error(f"Ошибка при обработке Google Sheet: {error}", exc_info=True)
        return None
